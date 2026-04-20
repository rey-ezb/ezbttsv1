import unittest
from datetime import datetime

from openpyxl import Workbook

from demand_planning_app.inventory_sheet import (
    INVENTORY_HISTORY_COLUMNS,
    build_inventory_history_from_workbook,
)


def _build_base_workbook() -> Workbook:
    workbook = Workbook()
    historical = workbook.active
    historical.title = "Inventory Historical Data"
    updated = workbook.create_sheet("Updated Report")
    workbook.create_sheet("ME Inventory Totals")
    workbook.create_sheet("Daily Inventory Report ")
    return workbook


def _set_historical_headers(sheet) -> None:
    sheet["A1"] = "Date"
    sheet["B1"] = "TikTok "
    sheet["T1"] = "Amazon"

    sheet["B2"] = "Birria"
    sheet["E2"] = "Pozole"
    sheet["H2"] = "Tinga"
    sheet["K2"] = "Brine"
    sheet["N2"] = "Variety Pack"
    sheet["Q2"] = "Pozole Verde"
    sheet["T2"] = "Birria"

    for start in ("B", "E", "H", "K", "N", "Q", "T"):
        row = 3
        col = sheet[f"{start}{row}"].column
        sheet.cell(row=row, column=col, value="In Transit")
        sheet.cell(row=row, column=col + 1, value="On Hand")
        sheet.cell(row=row, column=col + 2, value="$")


def _set_updated_headers(sheet, as_of: str) -> None:
    sheet["A1"] = "EZ Bombs Inventory Report"
    sheet["A2"] = f"As at {as_of}"
    sheet["C5"] = "Birria"
    sheet["E5"] = "Pozole"
    sheet["G5"] = "Tinga"
    sheet["I5"] = "Brine Bomb"
    sheet["K5"] = "Pozole Verde"
    sheet["M5"] = "Variety Pack"
    for column in ("C", "E", "G", "I", "K", "M"):
        base = sheet[f"{column}6"].column
        sheet.cell(row=6, column=base, value="In-Transit")
        sheet.cell(row=6, column=base + 1, value="On-hand")
    sheet["B8"] = "     TikTok"


class InventorySheetTests(unittest.TestCase):
    def test_build_inventory_history_from_workbook_keeps_tiktok_supply_only(self) -> None:
        workbook = _build_base_workbook()
        historical = workbook["Inventory Historical Data"]
        updated = workbook["Updated Report"]
        _set_historical_headers(historical)
        _set_updated_headers(updated, "04/13/2026")

        historical["A4"] = datetime(2026, 4, 14)
        historical["B4"] = 12
        historical["C4"] = 34
        historical["D4"] = 999
        historical["E4"] = 0
        historical["F4"] = 21
        historical["G4"] = 555
        historical["T4"] = 77
        historical["U4"] = 88

        historical["A5"] = datetime(2026, 4, 15)

        inventory = build_inventory_history_from_workbook(workbook, channel="TikTok")

        self.assertEqual(list(inventory.columns), INVENTORY_HISTORY_COLUMNS)
        self.assertEqual(len(inventory), 2)
        self.assertEqual(set(inventory["product_name"]), {"Birria Bomb 2-Pack", "Pozole Bomb 2-Pack"})
        self.assertTrue((inventory["platform"] == "TikTok").all())
        birria = inventory.loc[inventory["product_name"].eq("Birria Bomb 2-Pack")].iloc[0]
        self.assertEqual(float(birria["in_transit"]), 12.0)
        self.assertEqual(float(birria["on_hand"]), 34.0)

    def test_build_inventory_history_from_workbook_appends_newer_updated_snapshot(self) -> None:
        workbook = _build_base_workbook()
        historical = workbook["Inventory Historical Data"]
        updated = workbook["Updated Report"]
        _set_historical_headers(historical)
        _set_updated_headers(updated, "04/15/2026")

        historical["A4"] = datetime(2026, 4, 14)
        historical["B4"] = 0
        historical["C4"] = 100
        historical["E4"] = 0
        historical["F4"] = 50
        historical["H4"] = 0
        historical["I4"] = 25

        updated["C8"] = 0
        updated["D8"] = 111
        updated["E8"] = 0
        updated["F8"] = 222
        updated["G8"] = 0
        updated["H8"] = 333
        updated["I8"] = 0
        updated["J8"] = 444
        updated["K8"] = 0
        updated["L8"] = 555
        updated["M8"] = 0
        updated["N8"] = 666

        inventory = build_inventory_history_from_workbook(workbook, channel="TikTok")

        self.assertEqual(inventory["date"].nunique(), 2)
        latest = inventory.loc[inventory["date"].eq(inventory["date"].max())]
        self.assertEqual(set(latest["product_name"]), {
            "Birria Bomb 2-Pack",
            "Pozole Bomb 2-Pack",
            "Tinga Bomb 2-Pack",
            "Brine Bomb",
            "Pozole Verde Bomb 2-Pack",
            "Variety Pack",
        })
        pozole_verde = latest.loc[latest["product_name"].eq("Pozole Verde Bomb 2-Pack")].iloc[0]
        self.assertEqual(float(pozole_verde["on_hand"]), 555.0)


if __name__ == "__main__":
    unittest.main()
