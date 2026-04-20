from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


INVENTORY_HISTORY_COLUMNS = ["platform", "date", "product_name", "on_hand", "in_transit"]

INVENTORY_PRODUCT_MAP = {
    "birria": "Birria Bomb 2-Pack",
    "pozole": "Pozole Bomb 2-Pack",
    "tinga": "Tinga Bomb 2-Pack",
    "brine": "Brine Bomb",
    "brine bomb": "Brine Bomb",
    "variety pack": "Variety Pack",
    "pozole verde": "Pozole Verde Bomb 2-Pack",
}

SUPPORTED_METRICS = {"on_hand", "in_transit"}


def inventory_export_url(source: str | Path) -> str:
    raw = str(source)
    if "docs.google.com/spreadsheets/d/" not in raw:
        return raw
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", raw)
    if not match:
        raise ValueError(f"Unsupported Google Sheets URL: {raw}")
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"


def load_inventory_workbook(source: str | Path) -> Workbook:
    raw = str(source)
    if raw.startswith("http://") or raw.startswith("https://"):
        response = requests.get(inventory_export_url(raw), timeout=60)
        response.raise_for_status()
        return load_workbook(io.BytesIO(response.content), data_only=True)
    return load_workbook(Path(raw), data_only=True)


def build_inventory_history_from_workbook(workbook: Workbook, *, channel: str = "TikTok") -> pd.DataFrame:
    historical = _parse_inventory_historical_sheet(workbook["Inventory Historical Data"], channel=channel)
    updated = _parse_updated_report_sheet(workbook["Updated Report"], channel=channel)
    if not updated.empty and (historical.empty or updated["date"].max() > historical["date"].max()):
        historical = pd.concat([historical, updated], ignore_index=True)
    if historical.empty:
        return pd.DataFrame(columns=INVENTORY_HISTORY_COLUMNS)
    historical = historical[(historical["on_hand"] != 0) | (historical["in_transit"] != 0)]
    return historical.sort_values(["date", "product_name"]).reset_index(drop=True)


def load_inventory_history(source: str | Path, *, channel: str = "TikTok") -> pd.DataFrame:
    workbook = load_inventory_workbook(source)
    return build_inventory_history_from_workbook(workbook, channel=channel)


def _parse_inventory_historical_sheet(sheet: Worksheet, *, channel: str) -> pd.DataFrame:
    channels = _carry_forward_header_values(sheet, row=1)
    products = _carry_forward_header_values(sheet, row=2)
    metrics = [_normalize_metric_name(sheet.cell(3, column).value) for column in range(1, sheet.max_column + 1)]

    rows: list[dict[str, Any]] = []
    for row_index in range(4, sheet.max_row + 1):
        raw_date = sheet.cell(row_index, 1).value
        if pd.isna(raw_date):
            continue
        for column in range(2, sheet.max_column + 1):
            if channels[column - 1] != channel:
                continue
            product_name = _canonical_inventory_product_name(products[column - 1])
            metric_name = metrics[column - 1]
            if not product_name or metric_name not in SUPPORTED_METRICS:
                continue
            rows.append(
                {
                    "platform": channel,
                    "date": pd.Timestamp(raw_date).normalize(),
                    "product_name": product_name,
                    "metric": metric_name,
                    "value": _to_float(sheet.cell(row_index, column).value),
                }
            )
    return _pivot_inventory_metric_rows(rows, channel=channel)


def _parse_updated_report_sheet(sheet: Worksheet, *, channel: str) -> pd.DataFrame:
    as_of = _parse_as_of_date(sheet["A2"].value)
    if as_of is None:
        return pd.DataFrame(columns=INVENTORY_HISTORY_COLUMNS)

    products = _carry_forward_header_values(sheet, row=5)
    metrics = [_normalize_metric_name(sheet.cell(6, column).value) for column in range(1, sheet.max_column + 1)]

    channel_row = None
    for row_index in range(7, sheet.max_row + 1):
        if _clean_label(sheet.cell(row_index, 2).value) == channel:
            channel_row = row_index
            break
    if channel_row is None:
        return pd.DataFrame(columns=INVENTORY_HISTORY_COLUMNS)

    rows: list[dict[str, Any]] = []
    for column in range(3, sheet.max_column + 1):
        product_name = _canonical_inventory_product_name(products[column - 1])
        metric_name = metrics[column - 1]
        if not product_name or metric_name not in SUPPORTED_METRICS:
            continue
        rows.append(
            {
                "platform": channel,
                "date": pd.Timestamp(as_of),
                "product_name": product_name,
                "metric": metric_name,
                "value": _to_float(sheet.cell(channel_row, column).value),
            }
        )
    return _pivot_inventory_metric_rows(rows, channel=channel)


def _pivot_inventory_metric_rows(rows: list[dict[str, Any]], *, channel: str) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=INVENTORY_HISTORY_COLUMNS)
    frame = pd.DataFrame(rows)
    wide = (
        frame.pivot_table(
            index=["platform", "date", "product_name"],
            columns="metric",
            values="value",
            aggfunc="first",
        )
        .reset_index()
        .fillna(0.0)
    )
    wide.columns.name = None
    for metric in SUPPORTED_METRICS:
        if metric not in wide.columns:
            wide[metric] = 0.0
    wide["platform"] = channel
    return wide[INVENTORY_HISTORY_COLUMNS]


def _carry_forward_header_values(sheet: Worksheet, *, row: int) -> list[str]:
    values: list[str] = []
    last_value = ""
    for column in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row, column).value
        if cell_value is not None:
            last_value = _clean_label(cell_value)
        values.append(last_value)
    return values


def _clean_label(value: Any) -> str:
    return str(value or "").strip()


def _normalize_metric_name(value: Any) -> str:
    label = _clean_label(value).lower().replace("-", " ").replace("$", "")
    label = re.sub(r"\s+", " ", label).strip()
    if label == "on hand":
        return "on_hand"
    if label == "in transit":
        return "in_transit"
    return label


def _canonical_inventory_product_name(value: Any) -> str | None:
    label = _clean_label(value).lower()
    return INVENTORY_PRODUCT_MAP.get(label)


def _parse_as_of_date(value: Any) -> pd.Timestamp | None:
    text = _clean_label(value)
    match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", text)
    if not match:
        return None
    return pd.Timestamp(match.group(1))


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    text = str(value).replace(",", "").strip()
    if not text:
        return 0.0
    return float(text)
